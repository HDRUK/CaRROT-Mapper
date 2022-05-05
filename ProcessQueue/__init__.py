import logging
import json
import azure.functions as func
from azure.storage.blob import BlobServiceClient
from io import BytesIO
import requests
import openpyxl
from datetime import datetime
import os
import csv
import psutil
import httpx
import asyncio

from requests.models import HTTPError
from collections import defaultdict

from shared_code import omop_helpers

# import memory_profiler
# root_logger = logging.getLogger()
# root_logger.handlers[0].setFormatter(logging.Formatter("%(name)s: %(message)s"))
# profiler_logstream = memory_profiler.LogFile('memory_profiler_logs', True)

logger = logging.getLogger("test_logger")

stream_handler = logging.StreamHandler()
stream_handler.setFormatter(
    logging.Formatter(
        fmt="%(asctime)s - %(levelname)s - %(message)s", datefmt="%d/%m/%Y " "%H:%M:%S"
    )
)
logger.addHandler(stream_handler)
logger.setLevel(logging.INFO)  # Set to logging.DEBUG to show the debug output


# Agreed vocabs that are accepted for lookup/conversion
# The Data Team decide what vocabs are accepted.
# Add more as necessary by appending the list
vocabs = [
    "ABMS",
    "ATC",
    "HCPCS",
    "HES Specialty",
    "ICD10",
    "ICD10CM",
    "ICD10PCS",
    "ICD9CM",
    "ICD9Proc",
    "LOINC",
    "NDC",
    "NUCC",
    "OMOP Extension",
    "OSM",
    "PHDSC",
    "Read",
    "RxNorm",
    "RxNorm Extension",
    "SNOMED",
    "SPL",
    "UCUM",
    "UK Biobank",
]

max_chars_for_get = 2000


# @memory_profiler.profile(stream=profiler_logstream)
def process_scan_report_sheet_table(sheet):
    """
    This function extracts the
    data into the format below.

    -- Example Table Sheet CSV --
    a,   frequency,          b, frequency
    apple,      20,     orange,         5
    banana,      3,   plantain,        50
    pear,       12,         '',        ''

    --

    -- output --
    [(a,    apple, 20),
     (a,   banana,  3),
     (a,     pear, 12),
     (b,   orange,  5),
     (b, plantain, 50)]
    --
    """
    logger.debug("Start process_scan_report_sheet_table")

    sheet.reset_dimensions()
    sheet.calculate_dimension(force=True)
    # Get header entries (skipping every second column which is just 'Frequency')
    # So headers = ['a', 'b']
    first_row = sheet[1]
    headers = [cell.value for cell in first_row[::2]]

    # Set up an empty defaultdict, and fill it with one entry per header (i.e. one
    # per column)
    # Append each entry's value with the tuple (value, frequency) so that we end up
    # with each entry containing one tuple per non-empty entry in the column.
    #
    # This will give us
    #
    # ordereddict({'a': [('apple', 20), ('banana', 3), ('pear', 12)],
    #              'b': [('orange', 5), ('plantain', 50)]})

    d = defaultdict(list)
    # Iterate over all rows beyond the header - use the number of headers*2 to
    # set the maximum column rather than relying on sheet.max_col as this is not
    # always reliably updated by Excel etc.
    for row in sheet.iter_rows(
        min_col=1,
        max_col=len(headers) * 2,
        min_row=2,
        max_row=sheet.max_row,
        values_only=True,
    ):
        # Set boolean to track whether we hit a blank row for early exit below.
        this_row_empty = True
        # Iterate across the pairs of cells in the row. If the pair is non-empty,
        # then add it to the relevant dict entry.
        for (header, cell, freq) in zip(headers, row[::2], row[1::2]):
            if cell != "" or freq != "":
                d[header].append((str(cell), freq))
                this_row_empty = False
        # This will trigger if we hit a row that is entirely empty. Short-circuit
        # to exit early here - this saves us from situations where sheet.max_row is
        # incorrectly set (too large)
        if this_row_empty:
            break

    logger.debug("Finish process_scan_report_sheet_table")
    return d


def default_zero(input):
    """
    Helper function that returns the input, replacing anything Falsey
    (such as Nones or empty strings) with 0.0.
    """
    return round(input if input else 0.0, 2)


def handle_max_chars(max_chars=None):
    if not max_chars:
        max_chars = (
            int(os.environ.get("PAGE_MAX_CHARS"))
            if os.environ.get("PAGE_MAX_CHARS")
            else 10000
        )
    return max_chars


def perform_chunking(entries_to_post):
    """
    This expects a list of dicts, and returns a list of lists of lists of dicts,
    where the maximum length of each list of dicts, under JSONification,
    is less than max_chars, and the length of each list of lists of dicts is chunk_size
    """
    max_chars = handle_max_chars()
    chunk_size = (
        int(os.environ.get("CHUNK_SIZE")) if os.environ.get("CHUNK_SIZE") else 6
    )

    chunked_entries_to_post = []
    this_page = []
    this_chunk = []
    page_no = 0
    for entry in entries_to_post:
        # If the current page won't be overfull, add the entry to the current page
        if len(json.dumps(this_page)) + len(json.dumps(entry)) < max_chars:
            this_page.append(entry)
        # Otherwise, this page should be added to the current chunk.
        else:
            this_chunk.append(this_page)
            page_no += 1
            # Now check for a full chunk. If full, then add this chunk to the list of chunks.
            if page_no % chunk_size == 0:
                # append the chunk to the list of chunks, then reset the chunk to empty
                chunked_entries_to_post.append(this_chunk)
                this_chunk = []
            # Now add the entry that would have over-filled the page.
            this_page = [entry]
    # After all entries are added, check for a half-filled page, and if present add it to the list of pages
    if this_page:
        this_chunk.append(this_page)
    # Similarly, if a chunk ends up half-filled, add it to thelist of chunks
    if this_chunk:
        chunked_entries_to_post.append(this_chunk)

    return chunked_entries_to_post


def paginate(entries, max_chars=None):
    """
    This expects a list of strings, and returns a list of lists of strings,
    where the maximum length of each list of strings, under JSONification,
    is less than max_chars
    """
    max_chars = handle_max_chars(max_chars)

    paginated_entries = []
    this_page = []
    for entry in entries:
        # If the current page won't be overfull, add the entry to the current page
        if len(json.dumps(this_page)) + len(json.dumps(entry)) < max_chars:
            this_page.append(entry)
        else:
            # Otherwise, this page should be added to the list of pages.
            paginated_entries.append(this_page)
            # Now add the entry that would have over-filled the page.
            this_page = [entry]

    # After all entries are added, check for a half-filled page, and if present add it to the list of pages
    if this_page:
        paginated_entries.append(this_page)

    return paginated_entries


# @memory_profiler.profile(stream=profiler_logstream)
def startup(msg):
    logger.info("Python queue trigger function processed a queue item.")
    logger.debug(f"RAM memory % used: {psutil.virtual_memory()}")
    # Set up ccom API parameters:
    api_url = os.environ.get("APP_URL") + "api/"
    headers = {
        "Content-type": "application/json",
        "charset": "utf-8",
        "Authorization": "Token {}".format(os.environ.get("AZ_FUNCTION_KEY")),
    }

    # Get message from queue
    message = {
        "id": msg.id,
        "body": msg.get_body().decode("utf-8"),
        "expiration_time": (
            msg.expiration_time.isoformat() if msg.expiration_time else None
        ),
        "insertion_time": (
            msg.insertion_time.isoformat() if msg.insertion_time else None
        ),
        "time_next_visible": (
            msg.time_next_visible.isoformat() if msg.time_next_visible else None
        ),
        "pop_receipt": msg.pop_receipt,
        "dequeue_count": msg.dequeue_count,
    }

    logger.info(f"message: {message}")
    # Grab message body from storage queues,
    # extract filenames for scan reports and dictionaries
    # print("body 1:", type(message["body"]), message["body"])
    body = json.loads(message["body"])
    # print("body 2:", type(body), body)
    scan_report_blob = body["scan_report_blob"]
    data_dictionary_blob = body["data_dictionary_blob"]

    logger.info(f"MESSAGE BODY >>> {body}")

    # If the message has been dequeued for a second time, then the upload has failed.
    # Patch the name of the dataset to make it clear that it has failed,
    # set the status to 'Upload Failed', and then stop.
    logger.info(f"dequeue_count {msg.dequeue_count}")
    scan_report_id = body["scan_report_id"]
    if msg.dequeue_count == 2:
        process_failure(api_url, scan_report_id, headers)

    if msg.dequeue_count > 1:
        raise Exception("dequeue_count > 1")

    # Otherwise, this must be the first time we've seen this message. Proceed.
    return api_url, headers, scan_report_blob, data_dictionary_blob, scan_report_id


def process_failure(api_url, scan_report_id, headers):
    scan_report_fetched_data = requests.get(
        url=f"{api_url}scanreports/{scan_report_id}/",
        headers=headers,
    )

    scan_report_fetched_data = json.loads(
        scan_report_fetched_data.content.decode("utf-8")
    )

    json_data = json.dumps({"status": "UPFAILE"})

    failure_response = requests.patch(
        url=f"{api_url}scanreports/{scan_report_id}/", data=json_data, headers=headers
    )


def flatten(arr):
    """
    This expects a list of lists and returns a flattened list
    """
    newArr = [item for sublist in arr for item in sublist]
    return newArr


def reuse_existing_field_concepts(new_fields_map, content_type, api_url, headers):
    """
    This expects a dict of field names to ids which have been generated in a newly uploaded
    scanreport, and content_type 15. It creates new concepts associated to any
    field that matches the name of an existing field with an associated concept.
    """
    logger.info(f"reuse_existing_field_concepts")
    # Gets all scan report concepts that are for the type field (or content type which should be field)
    get_field_concept_ids = requests.get(
        url=f"{api_url}scanreportconceptsfilter/?content_type="
        f"{content_type}&fields=id,object_id,concept",
        headers=headers,
    )
    # create dictionary that maps existing field ids to scan report concepts
    # from the list of existing scan report concepts
    existing_field_concept_ids = json.loads(
        get_field_concept_ids.content.decode("utf-8")
    )
    existing_field_id_to_concept_map = {
        str(element.get("object_id", None)): str(element.get("concept", None))
        for element in existing_field_concept_ids
    }
    logger.debug(
        f"field_id:concept_id for all existing fields with concepts: "
        f"{existing_field_id_to_concept_map}"
    )

    # print("FIELD TO CONCEPT MAP DICT", existing_field_id_to_concept_map)
    # creates a list of field ids from fields that already exist and have a concept
    existing_ids = list(existing_field_id_to_concept_map.keys())

    # paginate the field id's variable and field names from list of newly generated
    # fields so that get requests do not exceed character limit
    paginated_existing_ids = paginate(existing_ids, max_chars_for_get)
    paginated_new_field_names = paginate(list(new_fields_map.keys()), max_chars_for_get)
    # for each list in paginated ids, get scanreport fields that match any of the given
    # ids (those with an associated concept)
    existing_fields_filtered_by_id = []
    for ids in paginated_existing_ids:
        ids_to_get = ",".join(map(str, ids))

        get_field_tables = requests.get(
            url=f"{api_url}scanreportfields/?id__in={ids_to_get}&fields=id,"
            f"scan_report_table,name",
            headers=headers,
        )
        existing_fields_filtered_by_id.append(
            json.loads(get_field_tables.content.decode("utf-8"))
        )
    existing_fields_filtered_by_id = flatten(existing_fields_filtered_by_id)

    # for each list in paginated ids, get scanreport fields whose name matches any of
    # the newly generated names
    existing_fields_filtered_by_name = []
    for ids in paginated_new_field_names:
        ids_to_get = ",".join(map(str, ids))

        get_field_tables = requests.get(
            url=f"{api_url}scanreportfields/?name__in={ids_to_get}&fields=id,"
            f"scan_report_table",
            headers=headers,
        )
        existing_fields_filtered_by_name.append(
            json.loads(get_field_tables.content.decode("utf-8"))
        )
    existing_fields_filtered_by_name = flatten(existing_fields_filtered_by_name)

    # Combine the results of the two sets of GET requests to identify fields which
    # satisfy both criteria (id and name) and then store their details in
    # existing_field_details
    cofiltered_field_ids = set(
        field["id"] for field in existing_fields_filtered_by_id
    ).intersection(set(field["id"] for field in existing_fields_filtered_by_name))
    existing_fields_details = [
        field
        for field in existing_fields_filtered_by_id
        if field["id"] in cofiltered_field_ids
    ]

    # get table ids from fields and repeat the process
    table_ids = set([item["scan_report_table"] for item in existing_fields_details])
    paginated_table_ids = paginate(table_ids, max_chars_for_get)
    existing_tables_details = []
    for ids in paginated_table_ids:
        ids_to_get = ",".join(map(str, ids))

        get_field_tables = requests.get(
            url=f"{api_url}scanreporttables/?id__in={ids_to_get}&fields=id,"
            f"scan_report",
            headers=headers,
        )
        existing_tables_details.append(
            json.loads(get_field_tables.content.decode("utf-8"))
        )
    existing_tables_details = flatten(existing_tables_details)

    # get all scan reports to be used to filter values by only values that come from
    # active scan reports that are marked as 'Mapping Complete'
    get_scan_reports = requests.get(
        url=f"{api_url}scanreports/",
        headers=headers,
    )
    get_datasets = requests.get(
        url=f"{api_url}datasets/",
        headers=headers,
    )
    # get active scanreports and map them to fields. Remove any fields in archived
    # reports or not marked as 'Mapping Complete'
    active_srs = []
    for item in get_scan_reports.json():
        if item["hidden"] is False and item["status"] == "COMPLET":
            for ds in get_datasets.json():
                # Exclude scan reports if their parent_dataset is archived
                if ds["id"] == item["parent_dataset"] and ds["hidden"] is False:
                    active_srs.append(str(item["id"]))
    # active reports is list of report ids that belong to an active dataset, are not archived, and have the status
    # 'Mapping Complete'

    # map value id to active scan report
    table_id_to_active_scanreport_map = {
        str(element["id"]): str(element["scan_report"])
        for element in existing_tables_details
        if str(element["scan_report"]) in active_srs
    }
    existing_field_id_to_active_scanreport_map = {
        str(element["id"]): table_id_to_active_scanreport_map[
            str(element["scan_report_table"])
        ]
        for element in existing_fields_details
        if str(element["scan_report_table"]) in table_id_to_active_scanreport_map
    }
    # filter fields to only include fields that are from active scan reports
    existing_fields_details_in_active_sr = [
        item
        for item in existing_fields_details
        if str(item["id"]) in existing_field_id_to_active_scanreport_map
    ]
    # print("FILTERED FIELDS", fields)

    existing_mappings_to_consider = [
        {
            "name": field["name"],
            "concept": existing_field_id_to_concept_map[str(field["id"])],
            "id": field["id"],
        }
        for field in existing_fields_details_in_active_sr
    ]
    logger.debug(f"{existing_mappings_to_consider=}")

    existing_field_name_to_id_map = {}
    for name in list(new_fields_map.keys()):
        mappings_matching_field_name = [
            mapping
            for mapping in existing_mappings_to_consider
            if mapping["name"] == name
        ]
        target_concept_ids = set(
            [mapping["concept"] for mapping in mappings_matching_field_name]
        )
        target_field_id = set(
            [mapping["id"] for mapping in mappings_matching_field_name]
        )
        if len(target_concept_ids) == 1:
            existing_field_name_to_id_map[str(name)] = str(target_field_id.pop())

    # replace existing_field_name_to_id_map with field name to concept id map
    # field_name_to_concept_id_map = { element.key: existing_field_id_to_concept_map[int(element.value)] for element in field_name_to_id_map }

    logger.debug(f"{existing_field_name_to_id_map=}")
    concepts_to_post = []
    concept_response_content = []
    logger.debug(f"{new_fields_map=}")
    logger.debug(f"{existing_field_id_to_concept_map=}")
    # print("NAME IDS", new_fields_map.keys())

    for name, new_field_id in new_fields_map.items():
        try:
            existing_field_id = existing_field_name_to_id_map[name]
            concept_id = existing_field_id_to_concept_map[str(existing_field_id)]

            logger.info(
                f"Found existing field with id: {existing_field_id} with existing "
                f"concept mapping: {concept_id} which matches new field id: "
                f"{new_field_id}"
            )
            # Create ScanReportConcept entry for copying over the concept
            concept_entry = {
                "nlp_entity": None,
                "nlp_entity_type": None,
                "nlp_confidence": None,
                "nlp_vocabulary": None,
                "nlp_processed_string": None,
                "concept": concept_id,
                "object_id": new_field_id,
                "content_type": content_type,
                "creation_type": "R",
            }
            concepts_to_post.append(concept_entry)
        except KeyError:
            continue

    if concepts_to_post:
        paginated_concepts_to_post = paginate(concepts_to_post)
        concept_response = []
        for concepts_to_post_item in paginated_concepts_to_post:
            post_concept_response = requests.post(
                url=api_url + "scanreportconcepts/",
                headers=headers,
                data=json.dumps(concepts_to_post_item),
            )
            logger.info(
                f"CONCEPTS SAVE STATUS >>>"
                f"{post_concept_response.status_code} "
                f"{post_concept_response.reason}"
            )
            concept_response.append(
                json.loads(post_concept_response.content.decode("utf-8"))
            )
        concept_content = flatten(concept_response)

        concept_response_content += concept_content

        logger.info("POST concepts all finished in reuse_existing_field_concepts")


def reuse_existing_value_concepts(new_values_map, content_type, api_url, headers):
    """
    This expects a dict of value names to ids which have been generated in a newly uploaded scanreport and
    creates new concepts if any matching names are found with existing fields
    """
    logger.info("reuse_existing_value_concepts")
    # get all scan report concepts with the content type of values
    get_value_concept_ids = requests.get(
        url=f"{api_url}scanreportconceptsfilter/?content_type={content_type}&fields=object_id,concept",
        headers=headers,
    )
    # create dictionary that maps existing value ids to scan report concepts
    # from the list of existing scan report concepts
    existing_value_concept_ids = json.loads(
        get_value_concept_ids.content.decode("utf-8")
    )
    existing_value_id_to_concept_map = {
        str(element.get("object_id", None)): str(element.get("concept", None))
        for element in existing_value_concept_ids
    }

    new_paginated_field_ids = paginate(
        [value["scan_report_field"] for value in new_values_map], max_chars_for_get
    )
    logger.debug("new_paginated_field_ids")

    new_fields = []
    for ids in new_paginated_field_ids:
        ids_to_get = ",".join(map(str, ids))
        get_fields = requests.get(
            url=f"{api_url}scanreportfields/?id__in={ids_to_get}&fields=id,name",
            headers=headers,
        )
        new_fields.append(json.loads(get_fields.content.decode("utf-8")))
    new_fields = flatten(new_fields)
    logger.debug(f"fields of newly generated values: {new_fields}")

    new_fields_to_name_map = {str(field["id"]): field["name"] for field in new_fields}
    logger.debug(
        f"id:name of fields of newly generated values: " f"{new_fields_to_name_map}"
    )

    # TODO: Consider making this a tuple-dict like value_details_to_id_map?
    new_values_full_details = [
        {
            "name": value["value"],
            "description": value["value_description"],
            "field_name": new_fields_to_name_map[str(value["scan_report_field"])],
            "id": value["id"],
        }
        for value in new_values_map
    ]
    logger.debug(
        f"name, desc, field_name, id of newly-generated values: "
        f"{new_values_full_details}",
    )

    # create list of names of newly generated values
    new_values_names_list = list(set(value["value"] for value in new_values_map))
    logger.debug(f"newly generated values: {new_values_names_list}")

    # paginate list of value ids from existing values that have scanreport concepts and
    # use the list to get existing scanreport values that match the list any of the newly generated names

    paginated_existing_ids = paginate(
        [str(element.get("object_id", None)) for element in existing_value_concept_ids],
        max_chars_for_get,
    )
    logger.debug(f"paginated_existing_ids")

    paginated_new_value_names = paginate(new_values_names_list, max_chars_for_get)
    logger.debug(f"paginated_new_value_names")

    # for each list in paginated ids, get scanreport values that match any of the given
    # ids (those with an associated concept)
    existing_values_filtered_by_id = []
    for ids in paginated_existing_ids:
        ids_to_get = ",".join(map(str, ids))

        get_field_tables = requests.get(
            url=f"{api_url}scanreportvalues/?id__in={ids_to_get}&fields=id,value,scan_report_field,"
            f"value_description",
            headers=headers,
        )
        existing_values_filtered_by_id.append(
            json.loads(get_field_tables.content.decode("utf-8"))
        )
    logger.debug(f"existing_values_filtered_by_id")

    existing_values_filtered_by_id = flatten(existing_values_filtered_by_id)
    logger.debug(f"existing_values_filtered_by_id flattened")

    # for each list in paginated ids, get scanreport values whose name matches any of
    # the newly generated names
    existing_values_filtered_by_name = []
    for names in paginated_new_value_names:
        new_values_names = ",".join(map(str, names))

        get_field_tables = requests.get(
            url=f"{api_url}scanreportvalues/?value__in={new_values_names}&fields="
            f"id,value,scan_report_field,value_description",
            headers=headers,
        )
        existing_values_filtered_by_name.append(
            json.loads(get_field_tables.content.decode("utf-8"))
        )
    logger.debug(f"existing_values_filtered_by_name")

    existing_values_filtered_by_name = flatten(existing_values_filtered_by_name)
    logger.debug(f"existing_values_filtered_by_name flattened")

    # Combine the results of the two sets of GET requests to identify values which
    # satisfy both criteria (id and name) and then store their details in
    # existing_value_details
    cofiltered_value_ids = set(
        value["id"] for value in existing_values_filtered_by_id
    ).intersection(set(value["id"] for value in existing_values_filtered_by_name))
    existing_values_details = [
        value
        for value in existing_values_filtered_by_id
        if value["id"] in cofiltered_value_ids
    ]

    logger.debug(
        f"Details of existing values which have an associated concept and "
        f"match one of the new value names: {existing_values_details}"
    )

    # get field ids from values and use to get scan report fields
    field_ids = set([item["scan_report_field"] for item in existing_values_details])
    paginated_field_ids = paginate(field_ids, max_chars_for_get)
    existing_fields_details = []
    for ids in paginated_field_ids:
        ids_to_get = ",".join(map(str, ids))

        get_value_fields = requests.get(
            url=f"{api_url}scanreportfields/?id__in={ids_to_get}&fields=id,"
            f"name,scan_report_table",
            headers=headers,
        )
        existing_fields_details.append(
            json.loads(get_value_fields.content.decode("utf-8"))
        )
    existing_fields_details = flatten(existing_fields_details)
    existing_field_id_to_name_map = {
        str(field["id"]): field["name"] for field in existing_fields_details
    }
    logger.debug(f"{existing_field_id_to_name_map=}")

    # get table ids from fields and repeat the process
    table_ids = set([item["scan_report_table"] for item in existing_fields_details])
    paginated_table_ids = paginate(table_ids, max_chars_for_get)
    existing_tables_details = []
    for ids in paginated_table_ids:
        ids_to_get = ",".join(map(str, ids))

        get_field_tables = requests.get(
            url=f"{api_url}scanreporttables/?id__in={ids_to_get}&fields=id,"
            f"scan_report",
            headers=headers,
        )
        existing_tables_details.append(
            json.loads(get_field_tables.content.decode("utf-8"))
        )
    existing_tables_details = flatten(existing_tables_details)

    # get all scan reports to be used to filter values by only values that come from
    # active scan reports that are marked as 'Mapping Complete'
    get_scan_reports = requests.get(
        url=f"{api_url}scanreports/",
        headers=headers,
    )
    get_datasets = requests.get(
        url=f"{api_url}datasets/",
        headers=headers,
    )
    # get active scanreports and map them to fields. Remove any fields in archived
    # reports or not marked as 'Mapping Complete'
    active_srs = []
    for item in get_scan_reports.json():
        if item["hidden"] is False and item["status"] == "COMPLET":
            for ds in get_datasets.json():
                # Exclude scan reports if their parent_dataset is archived
                if ds["id"] == item["parent_dataset"] and ds["hidden"] is False:
                    active_srs.append(str(item["id"]))
    # active reports is list of report ids that belong to an active dataset, are not archived, and have the status
    # 'Mapping Complete'

    # map value id to active scan report
    table_id_to_active_scanreport_map = {
        str(element["id"]): str(element["scan_report"])
        for element in existing_tables_details
        if str(element["scan_report"]) in active_srs
    }
    field_id_to_active_scanreport_map = {
        str(element["id"]): table_id_to_active_scanreport_map[
            str(element["scan_report_table"])
        ]
        for element in existing_fields_details
        if str(element["scan_report_table"]) in table_id_to_active_scanreport_map
    }

    existing_value_id_to_active_scanreport_map = {
        str(element["id"]): field_id_to_active_scanreport_map[
            str(element["scan_report_field"])
        ]
        for element in existing_values_details
        if str(element["scan_report_field"]) in field_id_to_active_scanreport_map
    }
    logger.debug(f"{existing_value_id_to_active_scanreport_map=}")

    existing_values_details_in_active_sr = [
        item
        for item in existing_values_details
        if str(item["id"]) in existing_value_id_to_active_scanreport_map
    ]
    logger.debug(f"{existing_values_details_in_active_sr=}")

    # List of dicts, one dict per existing value in an active SR, with details of the
    # value and its field and concept
    existing_mappings_to_consider = [
        {
            "name": value["value"],
            "concept": existing_value_id_to_concept_map[str(value["id"])],
            "id": value["id"],
            "description": value["value_description"],
            "field_name": existing_field_id_to_name_map[
                str(value["scan_report_field"])
            ],
        }
        for value in existing_values_details_in_active_sr
    ]
    logger.debug(f"{existing_mappings_to_consider=}")

    value_details_to_id_map = {}
    for item in new_values_full_details:
        name = item["name"]
        description = item["description"]
        field_name = item["field_name"]
        mappings_matching_value_name = [
            mapping
            for mapping in existing_mappings_to_consider
            if mapping["name"] == name
            and mapping["description"] == description
            and mapping["field_name"] == field_name
        ]
        target_concept_ids = set(
            [mapping["concept"] for mapping in mappings_matching_value_name]
        )
        target_value_id = set(
            [mapping["id"] for mapping in mappings_matching_value_name]
        )
        if len(target_concept_ids) == 1:
            value_details_to_id_map[
                (str(name), str(description), str(field_name))
            ] = str(target_value_id.pop())

    concepts_to_post = []
    concept_response_content = []
    for new_value_detail in new_values_full_details:
        try:
            existing_value_id = value_details_to_id_map[
                (
                    str(new_value_detail["name"]),
                    str(new_value_detail["description"]),
                    str(new_value_detail["field_name"]),
                )
            ]
            # print("VALUE existing value id", existing_value_id)
            concept_id = existing_value_id_to_concept_map[str(existing_value_id)]
            # print("VALUE existing concept id", concept_id)
            new_value_id = str(new_value_detail["id"])
            logger.info(
                f"Found existing value with id: {existing_value_id} with existing "
                f"concept mapping: {concept_id} which matches new value id: "
                f"{new_value_id}"
            )
            # Create ScanReportConcept entry for copying over the concept
            concept_entry = {
                "nlp_entity": None,
                "nlp_entity_type": None,
                "nlp_confidence": None,
                "nlp_vocabulary": None,
                "nlp_processed_string": None,
                "concept": concept_id,
                "object_id": new_value_id,
                "content_type": content_type,
                "creation_type": "R",
            }
            concepts_to_post.append(concept_entry)
        except KeyError:
            continue

    if concepts_to_post:
        paginated_concepts_to_post = paginate(concepts_to_post)
        concept_response = []
        for concepts_to_post_item in paginated_concepts_to_post:
            post_concept_response = requests.post(
                url=api_url + "scanreportconcepts/",
                headers=headers,
                data=json.dumps(concepts_to_post_item),
            )
            logger.info(
                f"CONCEPTS SAVE STATUS >>> "
                f"{post_concept_response.status_code} "
                f"{post_concept_response.reason}"
            )
            concept_response.append(
                json.loads(post_concept_response.content.decode("utf-8"))
            )
        concept_content = flatten(concept_response)

        concept_response_content += concept_content

        logger.info("POST concepts all finished in reuse_existing_value_concepts")


def remove_BOM(intermediate):
    return [
        {key.replace("\ufeff", ""): value for key, value in d.items()}
        for d in intermediate
    ]


def process_three_item_dict(three_item_data):
    """
    Converts a list of dictionaries (each with keys 'csv_file_name', 'field_name' and
    'code') to a nested dictionary with indices 'csv_file_name', 'field_name' and
    internal value 'code'.

    [{'csv_file_name': 'table1', 'field_name': 'field1', 'value': 'value1', 'code':
    'code1'},
    {'csv_file_name': 'table1', 'field_name': 'field2', 'value': 'value2'},
    {'csv_file_name': 'table2', 'field_name': 'field2', 'value': 'value2', 'code':
    'code2'},
    {'csv_file_name': 'table3', 'field_name': 'field3', 'value': 'value3', 'code':
    'code3'}]
    ->
    {'table1': {'field1': 'value1', 'field2': 'value2'},
     'table2': {'field2': 'value2'},
     'table3': {'field3': 'value3}
    }
    """
    csv_file_names = set(row["csv_file_name"] for row in three_item_data)

    # Initialise the dictionary with the keys, and each value set to a blank dict()
    new_vocab_dictionary = dict.fromkeys(csv_file_names, dict())

    # Fill each subdict with the data from the input list
    for row in three_item_data:
        new_vocab_dictionary[row["csv_file_name"]][row["field_name"]] = row["code"]

    return new_vocab_dictionary


def process_four_item_dict(four_item_data):
    """
    Converts a list of dictionaries (each with keys 'csv_file_name', 'field_name' and
    'code' and 'value') to a nested dictionary with indices 'csv_file_name',
    'field_name', 'code', and internal value 'value'.

    [{'csv_file_name': 'table1', 'field_name': 'field1', 'value': 'value1', 'code':
    'code1'},
    {'csv_file_name': 'table1', 'field_name': 'field2', 'value': 'value2', 'code':
    'code2'},
    {'csv_file_name': 'table2', 'field_name': 'field2', 'value': 'value2', 'code':
    'code2'},
    {'csv_file_name': 'table2', 'field_name': 'field2', 'value': 'value3', 'code':
    'code3'},
    {'csv_file_name': 'table3', 'field_name': 'field3', 'value': 'value3', 'code':
    'code3'}]
    ->
    {'table1': {'field1': {'value1': 'code1'}, 'field2': {'value2': 'code2'}},
     'table2': {'field2': {'value2': 'code2', 'value3': 'code3'}},
     'table3': {'field3': {'value3': 'code3'}}
    }
    """
    csv_file_names = set(row["csv_file_name"] for row in four_item_data)

    # Initialise the dictionary with the keys, and each value set to a blank dict()
    new_data_dictionary = dict.fromkeys(csv_file_names, dict())

    for row in four_item_data:
        if row["field_name"] not in new_data_dictionary[row["csv_file_name"]]:
            new_data_dictionary[row["csv_file_name"]][row["field_name"]] = dict()
        new_data_dictionary[row["csv_file_name"]][row["field_name"]][row["code"]] = row[
            "value"
        ]

    return new_data_dictionary


# @memory_profiler.profile(stream=profiler_logstream)
def parse_blobs(scan_report_blob, data_dictionary_blob):
    logger.info("parse_blobs()")
    # Set Storage Account connection string
    blob_service_client = BlobServiceClient.from_connection_string(
        os.environ.get("STORAGE_CONN_STRING")
    )

    # Grab scan report data from blob
    streamdownloader = (
        blob_service_client.get_container_client("scan-reports")
        .get_blob_client(scan_report_blob)
        .download_blob()
    )
    scanreport_bytes = BytesIO(streamdownloader.readall())
    wb = openpyxl.load_workbook(
        scanreport_bytes, data_only=True, keep_links=False, read_only=True
    )

    # If dictionary is present, also download dictionary
    if data_dictionary_blob != "None":
        # Access data as StorageStreamerDownloader class
        # Decode and split the stream using csv.reader()
        dict_client = blob_service_client.get_container_client("data-dictionaries")
        blob_dict_client = dict_client.get_blob_client(data_dictionary_blob)

        # Grab all rows with 4 elements for use as value descriptions
        data_dictionary_intermediate = list(
            row
            for row in csv.DictReader(
                blob_dict_client.download_blob().readall().decode("utf-8").splitlines()
            )
            if row["value"] != ""
        )
        # Remove BOM from start of file if it's supplied.
        dictionary_data = remove_BOM(data_dictionary_intermediate)

        # Convert to nested dictionaries, with structure
        # {tables: {fields: {values: value description}}}
        data_dictionary = process_four_item_dict(dictionary_data)

        # Grab all rows with 3 elements for use as possible vocabs
        vocab_dictionary_intermediate = list(
            row
            for row in csv.DictReader(
                blob_dict_client.download_blob().readall().decode("utf-8").splitlines()
            )
            if row["value"] == ""
        )
        vocab_data = remove_BOM(vocab_dictionary_intermediate)

        # Convert to nested dictionaries, with structure
        # {tables: {fields: vocab}}
        vocab_dictionary = process_three_item_dict(vocab_data)

    else:
        data_dictionary = None
        vocab_dictionary = None

    return wb, data_dictionary, vocab_dictionary


# @memory_profiler.profile(stream=profiler_logstream)
def post_tables(fo_ws, api_url, scan_report_id, headers):
    # Get all the table names in the order they appear in the Field Overview page
    table_names = []
    # Iterate over cells in the first column, but because we're in ReadOnly mode we
    # can't do that in the simplest manner.
    fo_ws.reset_dimensions()
    fo_ws.calculate_dimension(force=True)
    for row in fo_ws.iter_rows(min_row=2, max_row=fo_ws.max_row):
        cell_value = row[0].value
        # Check value is both non-empty and not seen before
        if cell_value and cell_value not in table_names:
            table_names.append(cell_value)

    """
    For each table create a scan_report_table entry,
    Append entry to table_entries_to_post[] list,
    Create JSON array with all the entries,
    Send POST request to API with JSON as input,
    Save the response data(table IDs)
    """
    table_entries_to_post = []
    # print("Working on Scan Report >>>", scan_report_id)
    logger.debug(f"RAM memory % used: {psutil.virtual_memory()}")
    logger.info(f"TABLES NAMES >>> {table_names}")

    for table_name in table_names:
        # print("WORKING ON TABLE >>> ", table_name)

        # Truncate table names because sheet names are truncated to 31 characters in Excel
        short_table_name = table_name[:31]

        # Create ScanReportTable entry
        # Link to scan report using ID from the queue message
        table_entry = {
            "created_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
            "updated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
            "name": short_table_name,
            "scan_report": str(scan_report_id),
            "person_id": None,
            "birth_date": None,
            "measurement_date": None,
            "condition_date": None,
            "observation_date": None,
        }

        # print("SCAN REPORT TABLE ENTRY", table_entry)

        # Append to list
        table_entries_to_post.append(table_entry)

    logger.info("POST tables")
    # POST request to scanreporttables
    tables_response = requests.post(
        "{}scanreporttables/".format(api_url),
        data=json.dumps(table_entries_to_post),
        headers=headers,
    )

    logger.info("POST tables finished")

    logger.info(f"TABLE SAVE STATUS >>> {tables_response.status_code}")
    # Error on failure
    if tables_response.status_code != 201:
        process_failure(api_url, scan_report_id, headers)
        raise HTTPError(
            " ".join(
                [
                    "Error in table save:",
                    str(tables_response.status_code),
                    str(json.dumps(table_entries_to_post)),
                ]
            )
        )
    logger.debug(f"RAM memory % used: {psutil.virtual_memory()}")

    # Load the result of the post request,
    tables_content = json.loads(tables_response.content.decode("utf-8"))

    # Save the table ids that were generated from the POST method
    table_ids = [element["id"] for element in tables_content]

    logger.info(f"TABLE IDs {table_ids}")
    table_name_to_id_map = dict(zip(table_names, table_ids))
    return table_name_to_id_map


# @memory_profiler.profile(stream=profiler_logstream)
async def process_values_from_sheet(
    sheet,
    data_dictionary,
    vocab_dictionary,
    current_table_name,
    names_to_ids_dict,
    api_url,
    scan_report_id,
    headers,
):
    # print("WORKING ON", sheet.title)
    # Reset list for values
    value_entries_to_post = []
    # Get (col_name, value, frequency) for each field in the table
    fieldname_value_freq_dict = process_scan_report_sheet_table(sheet)

    """
    For every result of process_scan_report_sheet_table,
    Save the current name,value,frequency
    Create ScanReportValue entry,
    Append to value_entries_to_post[] list,
    Create JSON array with all the value entries, 
    Send POST request to API with JSON as input
    """
    for name, value_freq_tuples in fieldname_value_freq_dict.items():
        for full_value, frequency in value_freq_tuples:
            value = full_value[0:127]

            if not frequency:
                frequency = 0

            if data_dictionary is not None:
                # Look up value description. We use .get() to guard against
                # nonexistence in the dictionary without having to manually check. It
                # returns None if the value is not present
                table = data_dictionary.get(
                    str(current_table_name)
                )  # dict of fields in table
                if table:
                    field = data_dictionary[str(current_table_name)].get(
                        str(name)
                    )  # dict of values in field in table
                    if field:
                        val_desc = data_dictionary[str(current_table_name)][
                            str(name)
                        ].get(str(value))
                    else:
                        val_desc = None
                else:
                    val_desc = None

                # Grab data from the 'code' column in the data dictionary
                # 'code' can contain an ordinary value (e.g. Yes, No, Nurse, Doctor)
                # or it could contain one of our pre-defined vocab names
                # e.g. SNOMED, RxNorm, ICD9 etc.
                # We use .get() to guard against nonexistence in the dictionary
                # without having to manually check. It returns None if the value is
                # not present
                table = vocab_dictionary.get(
                    str(current_table_name)
                )  # dict of fields in table
                if table:
                    code = vocab_dictionary[str(current_table_name)].get(
                        str(name)
                    )  # dict of values, will default to None if field not found in table
                else:
                    code = None

                # If 'code' is in our vocab list, try and convert the ScanReportValue
                # (concept code) to conceptID
                # If there's a faulty concept code for the vocab, fail gracefully and
                # set concept_id to default (-1)
                if code in vocabs:
                    try:
                        concept_id = omop_helpers.get_concept_from_concept_code(
                            concept_code=value,
                            vocabulary_id=code,
                            no_source_concept=True,
                        )
                        concept_id = concept_id["concept_id"]
                    except:
                        concept_id = -1
                else:
                    concept_id = -1

            else:
                val_desc = None
                concept_id = -1

            # Create ScanReportValue entry
            # We temporarily utilise the redundant 'conceptID' field in ScanReportValue
            # to save any looked up conceptIDs in the previous block of code.
            # The conceptID will be cleared later
            scan_report_value_entry = {
                "created_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                "updated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                "value": value,
                "frequency": int(frequency),
                "conceptID": concept_id,
                "value_description": val_desc,
                "scan_report_field": names_to_ids_dict[name],
            }

            # Append to list
            value_entries_to_post.append(scan_report_value_entry)

    logger.info(
        f"POST {len(value_entries_to_post)} values to table {current_table_name}"
    )
    logger.debug(f"RAM memory % used: {psutil.virtual_memory()}")
    chunked_value_entries_to_post = perform_chunking(value_entries_to_post)
    values_response_content = []
    logger.debug(f"chunked values list len: {len(chunked_value_entries_to_post)}")
    timeout = httpx.Timeout(60.0, connect=30.0)

    page_count = 0
    for chunk in chunked_value_entries_to_post:
        async with httpx.AsyncClient(timeout=timeout) as client:
            tasks = []
            page_lengths = []
            for page in chunk:
                # POST value_entries_to_post to ScanReportValues model
                tasks.append(
                    asyncio.ensure_future(
                        client.post(
                            url="{}scanreportvalues/".format(api_url),
                            data=json.dumps(page),
                            headers=headers,
                        )
                    )
                )
                page_lengths.append(len(page))
                page_count += 1

            values_responses = await asyncio.gather(*tasks)

        for i, values_response in enumerate(values_responses):
            logger.info(
                f"VALUES SAVE STATUSES >>> {values_response.status_code} "
                f"{values_response.reason_phrase} {page_lengths[i]}"
            )

            if values_response.status_code != 201:
                process_failure(api_url, scan_report_id, headers)
                raise HTTPError(
                    " ".join(
                        [
                            "Error in values save:",
                            str(values_response.status_code),
                            str(json.dumps(page)),
                        ]
                    )
                )

            values_response_content += json.loads(
                values_response.content.decode("utf-8")
            )

    logger.info("POST values all finished")
    logger.debug(f"RAM memory % used: {psutil.virtual_memory()}")
    # Process conceptIDs in ScanReportValues
    # GET values where the conceptID != -1 (i.e. we've converted a concept code to conceptID in the previous code)
    logger.debug("GET posted values")
    get_ids_of_posted_values = requests.get(
        url=api_url + "scanreportvaluepks/?scan_report=" + str(scan_report_id),
        headers=headers,
    )
    logger.debug("GET posted values finished")

    ids_of_posted_values = json.loads(get_ids_of_posted_values.content.decode("utf-8"))

    # Create a list for a bulk data upload to the ScanReportConcept model

    concept_id_data = [
        {
            "nlp_entity": None,
            "nlp_entity_type": None,
            "nlp_confidence": None,
            "nlp_vocabulary": None,
            "nlp_processed_string": None,
            "concept": concept["conceptID"],
            "object_id": concept["id"],
            # TODO: we should query this value from the API
            # - via ORM it would be ContentType.objects.get(model='scanreportvalue').id,
            # but that's not available from an Azure Function.
            "content_type": 17,
            "creation_type": "V",
        }
        for concept in ids_of_posted_values
    ]

    logger.info(f"POST {len(concept_id_data)} concepts")

    paginated_concept_id_data = paginate(concept_id_data)

    concepts_response_content = []

    for page in paginated_concept_id_data:

        # POST the ScanReportConcept data to the model
        concepts_response = requests.post(
            url=api_url + "scanreportconcepts/",
            headers=headers,
            data=json.dumps(page),
        )

        logger.info(
            f"CONCEPT SAVE STATUS >>> "
            f"{concepts_response.status_code} "
            f"{concepts_response.reason}"
        )
        if concepts_response.status_code != 201:
            process_failure(api_url, scan_report_id, headers)
            raise HTTPError(
                " ".join(
                    [
                        "Error in concept save:",
                        str(concepts_response.status_code),
                        str(json.dumps(page)),
                    ]
                )
            )

        concepts_content = json.loads(concepts_response.content.decode("utf-8"))
        concepts_response_content += concepts_content

    logger.info("POST concepts all finished")
    logger.debug(f"RAM memory % used: {psutil.virtual_memory()}")
    # Update ScanReportValue to remove any data added to the conceptID field
    # conceptID field only used temporarily to hold the converted concept code -> conceptID
    # Now the conceptID is saved to the correct model (ScanReportConcept) there's no
    # need for the concept ID to also be saved to ScanReportValue::conceptID

    # Reset conceptID to -1 (default). This doesn't need pagination because it's a
    # loop over all relevant fields anyway
    put_update_json = json.dumps({"conceptID": -1})

    logger.info(f"PATCH {len(ids_of_posted_values)} values")

    for concept in ids_of_posted_values:
        logger.debug("PATCH value")
        value_response = requests.patch(
            url=api_url + "scanreportvalues/" + str(concept["id"]) + "/",
            headers=headers,
            data=put_update_json,
        )
        # print("PATCH value finished", datetime.utcnow().strftime("%H:%M:%S.%fZ"))
        if value_response.status_code != 200:
            process_failure(api_url, scan_report_id, headers)
            raise HTTPError(
                " ".join(
                    [
                        "Error in value save:",
                        str(value_response.status_code),
                        str(put_update_json),
                    ]
                )
            )

    logger.info("PATCH values finished")
    reuse_existing_field_concepts(names_to_ids_dict, 15, api_url, headers)
    reuse_existing_value_concepts(values_response_content, 17, api_url, headers)
    logger.debug(f"RAM memory % used: {psutil.virtual_memory()}")


def post_field_entries(field_entries_to_post, api_url, scan_report_id, headers):
    paginated_field_entries_to_post = paginate(field_entries_to_post)
    fields_response_content = []
    # POST Fields
    for page in paginated_field_entries_to_post:
        fields_response = requests.post(
            "{}scanreportfields/".format(api_url),
            data=json.dumps(page),
            headers=headers,
        )
        # print('dumped:', json.dumps(page))
        logger.info(
            f"FIELDS SAVE STATUS >>> {fields_response.status_code} "
            f"{fields_response.reason} {len(page)}"
        )

        if fields_response.status_code != 201:
            process_failure(api_url, scan_report_id, headers)
            raise HTTPError(
                " ".join(
                    [
                        "Error in fields save:",
                        str(fields_response.status_code),
                        str(json.dumps(page)),
                    ]
                )
            )

        fields_response_content += json.loads(fields_response.content.decode("utf-8"))

    logger.info("POST fields all finished")
    return fields_response_content


def main(msg: func.QueueMessage):
    api_url, headers, scan_report_blob, data_dictionary_blob, scan_report_id = startup(
        msg
    )
    # Set the status to 'Upload in progress'
    status_in_progress_response = requests.patch(
        url=f"{api_url}scanreports/{scan_report_id}/",
        data=json.dumps({"status": "UPINPRO"}),
        headers=headers,
    )

    wb, data_dictionary, vocab_dictionary = parse_blobs(
        scan_report_blob, data_dictionary_blob
    )
    # Get the first sheet 'Field Overview',
    # to populate ScanReportTable & ScanReportField models
    fo_ws = wb.worksheets[0]

    table_name_to_id_map = post_tables(fo_ws, api_url, scan_report_id, headers)

    """
    POST fields per table:
    For each row in Field Overview create an entry for scan_report_field,
    Empty row signifies end of fields in a table
    Append field entry to field_entries_to_post[] list,
    Create JSON array with all the field entries, 
    Send POST request to API with JSON as input,
    Save the response data(field ids,field names) in a dictionary
    Set the current working sheet to be the same as the current table
    Post the values for that table
    """
    field_entries_to_post = []

    # Loop over all rows in Field Overview sheet.
    # For sheets past the first two in the Scan Report
    # i.e. all 'data' sheets that are not Field Overview and Table Overview
    logger.info("Start fields loop")
    previous_row_value = None
    for i, row in enumerate(fo_ws.iter_rows(min_row=2, max_row=fo_ws.max_row), start=2):
        # Guard against unnecessary rows beyond the last true row with contents
        if (previous_row_value is None or previous_row_value == "") and (
            row[0].value is None or row[0].value == ""
        ):
            break
        previous_row_value = row[0].value

        if row[0].value != "" and row[0].value is not None:
            current_table_name = row[0].value
            # Create ScanReportField entry
            field_entry = {
                "scan_report_table": table_name_to_id_map[current_table_name],
                "created_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                "updated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                "name": str(row[1].value),
                "description_column": str(row[2].value),
                "type_column": str(row[3].value),
                "max_length": row[4].value,
                "nrows": row[5].value,
                "nrows_checked": row[6].value,
                "fraction_empty": round(default_zero(row[7].value), 2),
                "nunique_values": row[8].value,
                "fraction_unique": round(default_zero(row[9].value), 2),
                "ignore_column": None,
                # "is_patient_id": False,
                # "is_ignore": False,
                # "pass_from_source": True,
                # "classification_system": str(row[11].value),
                # "concept_id": -1,
                # "field_description": None,
            }
            # Append each entry to a list
            field_entries_to_post.append(field_entry)

        else:
            # This is the scenario where the line is empty, so we're at the end of
            # the table. Don't add a field entry, but process all those so far.
            # print("scan_report_field_entries >>>", field_entries_to_post)

            # POST fields in this table
            logger.info(
                f"POST {len(field_entries_to_post)} fields to table "
                f"{current_table_name}"
            )
            logger.debug(f"RAM memory % used: {psutil.virtual_memory()}")

            fields_response_content = post_field_entries(
                field_entries_to_post, api_url, scan_report_id, headers
            )
            field_entries_to_post = []

            # Create a dictionary with field names and field ids from the response
            # as key value pairs
            # e.g ("Field Name": Field ID)
            names_to_ids_dict = {
                str(element.get("name", None)): str(element.get("id", None))
                for element in fields_response_content
            }

            # print("Dictionary id:name", names_to_ids_dict)

            if current_table_name not in wb.sheetnames:
                process_failure(api_url, scan_report_id, headers)
                raise ValueError(
                    f"Attempting to access sheet '{current_table_name}'"
                    f" in scan report, but no such sheet exists."
                )

            # Go to Table sheet to process all the values from the sheet
            sheet = wb[current_table_name]
            asyncio.run(
                process_values_from_sheet(
                    sheet,
                    data_dictionary,
                    vocab_dictionary,
                    current_table_name,
                    names_to_ids_dict,
                    api_url,
                    scan_report_id,
                    headers,
                )
            )
    logger.info("All tables completed. Now set status to 'Upload Complete'")
    # Set the status to 'Upload Complete'
    status_complete_response = requests.patch(
        url=f"{api_url}scanreports/{scan_report_id}/",
        data=json.dumps({"status": "UPCOMPL"}),
        headers=headers,
    )
    logger.info("Successfully set status to 'Upload Complete'")
    wb.close()
    logger.info("Workbook successfully closed")
